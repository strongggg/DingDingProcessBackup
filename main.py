import requests 
import re
import json
import openpyxl
import datetime
import os

banner = """


  _____                             ____             _                
 |  __ \                           |  _ \           | |               
 | |__) | __ ___   ___ ___  ___ ___| |_) | __ _  ___| | ___   _ _ __  
 |  ___/ '__/ _ \ / __/ _ \/ __/ __|  _ < / _` |/ __| |/ / | | | '_ \ 
 | |   | | | (_) | (_|  __/\__ \__ \ |_) | (_| | (__|   <| |_| | |_) |
 |_|   |_|  \___/ \___\___||___/___/____/ \__,_|\___|_|\_\\__,_| .__/ 
                                                               | |    
                                                               |_|    

                                                            v1.0
    
 
    用于钉钉审批模板备份
    输入Cookie，输出审批备份

"""


def parsenameandrole(result):
    dic = {}
    username = re.findall(r'\"userName\":\"([\s\S]+?)\"',result)
    label = re.findall(r'\"labelNames\":\"([\s\S]+?)\"',result)
    try:
        #processcode = re.search(r'\"processCode\":\"([\s\S]+?)\"',result).group(1)

        usernamelist = []
        for i in username:
            if i not in usernamelist:
                usernamelist.append(i)
        labellist = []
        for i in label:
            if i not in labellist:
                labellist.append(i)
        #print(usernamelist)
        #print(labellist)
        #print(processcode)
        usernamestr = str(usernamelist).replace('\'','').replace('[','').replace(']','').replace(' ','')
        labelstr = str(labellist).replace('\'','').replace('[','').replace(']','').replace(' ','')
        dic['usernamestr'] = usernamestr
        dic['labelstr'] = labelstr
        
    except Exception as e:
        print(e)

    return dic


def main(cookie):
    global resultlist
    csrftoken = re.search(r'_csrf_token_=(\d+)',cookie).group(1)
    print(csrftoken)
    headers = {
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:71.0) Gecko/20100101 Firefox/71.0',
        'Accept':'application/json, text/javascript, */*; q=0.01',
        'Accept-Language':'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
        'Content-Type':'application/x-www-form-urlencoded; charset=UTF-8',
        'Connection':'close',
        '_csrf_token_':csrftoken,
        'Cookie':cookie,
    }

    flowlisturl = "https://aflow.dingtalk.com/dingtalk/web/query/process/getMgrProcessList.json"

    data = {'locale':'zh-cn','isNeedFormContent':'false'}
    r = requests.post(flowlisturl,headers=headers,data=data)

    rjson = json.loads(r.text)
    #print(rjson)

    if rjson['httpStatus'] == '200':
        print('审批列表请求成功')
        data = rjson['data']
        processdirlist = data['sortedDirProcessList']
        for item in processdirlist:
            dirName = item['dirName']
            processformlist = item['sortedProcessAndFormVoList']
            for eachpro in processformlist:
                dic = {}
                dic['dirName'] = dirName
                dic['managerNicks'] = ','.join(eachpro['managerNicks'])
                dic['visibleSummaryText'] = eachpro['visibleSummaryText']
                dic['processCode'] = eachpro['processCode']
                dic['flowTitle'] = eachpro['flowTitle']
                dic['processStatus'] = eachpro['processStatus']
                #dic['gmtModified'] = eachpro['gmtModified']
                if dic['processStatus'] == 'PUBLISHED':
                    eachprourl = "https://aflow.dingtalk.com/dingtalk/web/query/form/getFormAndProcessConfig.json"
                    postdata = {
                        'processCode':eachpro['processCode'],
                        'appType':'0',
                        'tag':'',
                    }
                    r2 = requests.post(eachprourl,headers=headers,data=postdata)
                    r2json = json.loads(r2.text)
                    if r2json['httpStatus'] == '200':
                        print('['+eachpro['flowTitle']+']'+'详细模板请求成功')
                        r2data = r2json['data']
                        dic['content'] = r2data['formVo']['content']
                        dic['processConfig'] = r2data['processConfig']
                        d = parsenameandrole(r2data['processConfig'])
                        print(d)
                        dic['username'] = d['usernamestr']
                        dic['role'] = d['labelstr']
                        try:
                            dic['processguide'] = re.search(r'\"content\":\"([\s\S]+?)\"',r2data['formVo']['content']).group(1)
                        except Exception as e:
                            dic['processguide'] = ''
                            print(e)
                        dic['modifierName'] = r2data['modifierName']
                        dic['modifierTime'] = r2data['modifierTime']
                        resultlist.append(dic)


if __name__ == '__main__':
    print(banner)
    print()
    resultlist = []
    cookie = input("Cookie:\n")
    today = datetime.date.today().strftime("%Y%m%d")
    if not os.path.exists(today):
        os.makedirs(today)
    
    main(cookie)
    try:
        with open(today+'/resulttmp_'+ today +'.txt','w',encoding='utf8') as f:
            f.write(str(resultlist))
    except Exception as e:
        print(e)

    print('开始写入json文件')
    try:
        with open(today+'/result_'+ today +'.json','w',encoding='utf8') as f:
            json.dump(resultlist,f)
        print('写入json文件成功')
    except Exception as e:
        print(e)
    '''
    with open(today+'/result_'+ today +'.json','r',encoding='utf8') as f:
        resultlist = json.loads(f.read())
    '''

    print('开始写入excel文件')
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(index=0,title=today)
    row0 = ['ID','板块','审批名称','管理员','可见范围','流程状态','流程号','涉及人员姓名','涉及角色','审批指引','修改时间','修改人','表单内容','流程设计']
    for col in range(1,len(row0)+1,1):
        ws.cell(1,col,row0[col-1])
    for index,item in enumerate(resultlist):
        row = index + 2
        ws.cell(row,1,index)
        ws.cell(row,2,item['dirName'])
        ws.cell(row,3,item['flowTitle'])
        ws.cell(row,4,item['managerNicks'])
        ws.cell(row,5,item['visibleSummaryText'])
        ws.cell(row,6,item['processStatus'])
        ws.cell(row,7,item['processCode'])
        ws.cell(row,8,item['username'])
        ws.cell(row,9,item['role'])
        ws.cell(row,10,item['processguide'])
        ws.cell(row,11,item['modifierTime'])
        ws.cell(row,12,item['modifierName'])
        ws.cell(row,13,item['content'])
        ws.cell(row,14,item['processConfig'])
        #ws = wb["Sheet"]
        #wb.remove(ws)
    wb.save(today+'/result_'+ today +'.xlsx')
    print('写入excel文件成功')
    print('Done All !!!')

